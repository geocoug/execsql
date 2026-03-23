from __future__ import annotations

"""
XLS and XLSX spreadsheet export for execsql.

Provides :class:`XlsFile` (writes ``.xls`` via ``xlwt``) and
:class:`XlsxFile` (writes ``.xlsx`` via ``openpyxl``), both used by the
EXPORT metacommand.  Requires the ``execsql2[excel]`` extras.
"""

import datetime
import os
import re
from typing import Any, Optional, List

import execsql.state as _state
from execsql.exceptions import XlsFileError, XlsxFileError


class XlsFile:
    def __repr__(self) -> str:
        return "XlsFile()"

    class XlsLog:
        def __init__(self) -> None:
            self.log_msgs = []

        def write(self, msg: str) -> None:
            self.log_msgs.append(msg)

    def __init__(self) -> None:
        try:
            global xlrd
            import xlrd
        except:
            _state.fatal_error("The xlrd library is needed to read Excel (.xls) spreadsheets.")
        self.filename = None
        self.encoding = None
        self.wbk = None
        self.datemode = 0
        self.errlog = self.XlsLog()

    def open(self, filename: str, encoding: Optional[str] = None, read_only: bool = False) -> None:
        self.filename = filename
        self.encoding = encoding
        self.read_only = read_only
        if os.path.isfile(filename):
            # The 'read_only' argument is not used, but is present for compatibility with XlsxFile.open().
            self.wbk = xlrd.open_workbook(filename, logfile=self.errlog, encoding_override=self.encoding)
            self.datemode = self.wbk.datemode
        else:
            raise XlsFileError(f"There is no Excel file {self.filename}.")

    def sheetnames(self) -> Any:
        return self.wbk.sheets()

    def sheet_named(self, sheetname: Any) -> Any:
        # Return the sheet with the matching name.  If the name is actually an integer,
        # return that sheet number.
        if isinstance(sheetname, int):
            sheet_no = sheetname
        else:
            try:
                sheet_no = int(sheetname)
                if sheet_no < 1:
                    sheet_no = None
            except:
                sheet_no = None
        if sheet_no is None:
            sheet = self.wbk.sheet_by_name(sheetname)
        else:
            # User-specified sheet numbers should be 1-based; xlrd sheet indexes are 0-based
            sheet = self.wbk.sheet_by_index(max(0, sheet_no - 1))
        return sheet

    def sheet_data(self, sheetname: Any, junk_header_rows: int = 0) -> List:
        try:
            sheet = self.sheet_named(sheetname)
        except:
            raise XlsFileError(f"There is no Excel worksheet named {sheetname} in {self.filename}.")

        # Don't rely on sheet.ncols and sheet.nrows, because Excel will count columns
        # and rows that have ever been filled, even if they are now empty.  Base the column count
        # on the number of contiguous non-empty cells in the first row, and process the data up to nrows until
        # a row is entirely empty.
        def row_data(sheetrow, columns=None):
            cells = sheet.row_slice(sheetrow)
            if columns:
                d = [cells[c] for c in range(columns)]
            else:
                d = [cell for cell in cells]
            datarow = []
            for c in d:
                if c.ctype == 0:
                    # empty
                    datarow.append(None)
                elif c.ctype == 1:
                    # This might be a timestamp with time zone that xlrd treats as a string.
                    try:
                        dt = _state.DT_TimestampTZ()._from_data(c.value)
                        datarow.append(dt)
                    except:
                        datarow.append(c.value)
                elif c.ctype == 2:
                    # float, but maybe should be int
                    if c.value - int(c.value) == 0:
                        datarow.append(int(c.value))
                    else:
                        datarow.append(c.value)
                elif c.ctype == 3:
                    # date
                    dt = xlrd.xldate_as_tuple(c.value, self.datemode)
                    # Convert to time or datetime
                    if not any(dt[:3]):
                        # No date values
                        datarow.append(datetime.time(*dt[3:]))
                    else:
                        datarow.append(datetime.datetime(*dt))
                elif c.ctype == 4:
                    # Boolean
                    datarow.append(bool(c.value))
                elif c.ctype == 5:
                    # Error code
                    datarow.append(xlrd.error_text_from_code(c.value))
                elif c.ctype == 6:
                    # blank
                    datarow.append(None)
                else:
                    datarow.append(c.value)
            return datarow

        hdr_row = row_data(junk_header_rows)
        ncols = 0
        for c in range(len(hdr_row)):
            if not hdr_row[c]:
                break
            ncols += 1
        sheet_data = []
        for r in range(junk_header_rows, sheet.nrows - junk_header_rows):
            datarow = row_data(r, ncols)
            if datarow.count(None) == len(datarow):
                break
            sheet_data.append(datarow)
        return sheet_data


class XlsxFile:
    def __repr__(self) -> str:
        return "XlsxFile()"

    class XlsxLog:
        def __init__(self) -> None:
            self.log_msgs = []

        def write(self, msg: str) -> None:
            self.log_msgs.append(msg)

    def __init__(self) -> None:
        try:
            global openpyxl
            import openpyxl
        except:
            _state.fatal_error("The openpyxl library is needed to read Excel (.xlsx) spreadsheets.")
        self.filename = None
        self.encoding = None
        self.wbk = None
        self.read_only = False
        self.errlog = self.XlsxLog()

    def open(self, filename: str, encoding: Optional[str] = None, read_only: bool = False) -> None:
        self.filename = filename
        self.encoding = encoding
        self.read_only = read_only
        if os.path.isfile(filename):
            if read_only:
                self.wbk = openpyxl.load_workbook(filename, read_only=True)
            else:
                self.wbk = openpyxl.load_workbook(filename)
        else:
            raise XlsxFileError(f"There is no Excel file {self.filename}.")

    def close(self) -> None:
        if self.wbk is not None:
            self.wbk.close()
            self.wbk = None
            self.filename = None
            self.encoding = None

    def sheetnames(self) -> List[str]:
        return self.wbk.sheetnames

    def sheet_named(self, sheetname: Any) -> Any:
        # Return the sheet with the matching name.  If the name is actually an integer,
        # return that sheet number.
        if isinstance(sheetname, int):
            sheet_no = sheetname
        else:
            try:
                sheet_no = int(sheetname)
                if sheet_no < 1:
                    sheet_no = None
            except:
                sheet_no = None
        if sheet_no is not None:
            # User-specified sheet numbers should be 1-based
            sheet = self.wbk[self.wbk.sheetnames[sheet_no - 1]]
        else:
            sheet = self.wbk[sheetname]
        return sheet

    def sheet_data(self, sheetname: Any, junk_header_rows: int = 0) -> List:
        try:
            sheet = self.sheet_named(sheetname)
        except:
            raise XlsxFileError(f"There is no Excel worksheet named {sheetname} in {self.filename}.")
        # Don't rely on sheet.max_column and sheet.max_row, because Excel will count columns
        # and rows that have ever been filled, even if they are now empty.  Base the column count
        # on the number of contiguous non-empty cells in the first row, and process the data up to nrows until
        # a row is entirely empty.
        # Get the header row, skipping junk rows
        rowsrc = sheet.iter_rows(max_row=junk_header_rows + 1, values_only=True)
        for hdr_row in rowsrc:
            pass
        # Get the number of columns
        ncols = 0
        for c in range(len(hdr_row)):
            if not hdr_row[c]:
                break
            ncols += 1
        # Get all the data rows
        sheet_data = []
        rowsrc = sheet.iter_rows(min_row=junk_header_rows + 1, values_only=True)
        for r in rowsrc:
            if not any(r):
                break
            sheet_data.append(list(r))
        for r in range(len(sheet_data)):
            rd = sheet_data[r]
            for c in range(len(rd)):
                if isinstance(rd[c], str):
                    if rd[c] == "=FALSE()":
                        rd[c] = False
                    elif rd[c] == "=TRUE()":
                        rd[c] = True
        return sheet_data
