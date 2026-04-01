from __future__ import annotations

"""
XLSX (Excel Open XML) export for execsql.

Provides :func:`write_query_to_xlsx` (single-sheet export) and
:func:`write_queries_to_xlsx` (multi-sheet export).  Requires the
``openpyxl`` package (``execsql2[excel]``).
"""

import datetime
import getpass
import os
from pathlib import Path
from typing import Any

import execsql.state as _state
from execsql.exceptions import ErrInfo
from execsql.exporters.base import ExportRecord
from execsql.exporters.pretty import prettyprint_query
from execsql.script import current_script_line
from execsql.utils.errors import exception_desc, fatal_error
from execsql.utils.fileio import filewriter_close
from execsql.utils.strings import unquoted

__all__ = ["write_query_to_xlsx", "write_queries_to_xlsx"]

# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _require_openpyxl() -> Any:
    """Import and return the openpyxl module, raising a fatal error if absent."""
    try:
        import openpyxl

        return openpyxl
    except ImportError:
        fatal_error("The openpyxl library is needed to write Excel (.xlsx) spreadsheets (install execsql2[excel]).")


def _cell_value(item: Any) -> Any:
    """Return a value suitable for writing directly to an openpyxl cell.

    openpyxl natively handles int, float, bool, str, datetime.datetime,
    datetime.date, and None.  datetime.time is converted to a string because
    openpyxl does not have a native time-only cell type.
    """
    if item is None:
        return None
    if isinstance(item, bool):
        # bool must be checked before int — bool is a subclass of int.
        return item
    if isinstance(item, int | float):
        return item
    if isinstance(item, datetime.datetime):
        return item
    if isinstance(item, datetime.date):
        return item
    if isinstance(item, datetime.time):
        # openpyxl has no native time-only type; store as HH:MM:SS string.
        return item.strftime("%H:%M:%S")
    return str(item)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def write_query_to_xlsx(
    select_stmt: str,
    db: Any,
    outfile: str,
    append: bool = False,
    desc: str | None = None,
    sheetname: str | None = None,
) -> None:
    """Execute *select_stmt* and write the result to a single worksheet in an XLSX file.

    Args:
        select_stmt: SQL SELECT statement to execute.
        db: An execsql database adapter with a ``select_rowsource()`` method.
        outfile: Destination ``.xlsx`` file path.
        append: If ``True`` and *outfile* exists, add a new sheet to the
            existing workbook.  If ``False``, overwrite any existing file.
        desc: Optional human-readable description stored in the inventory sheet.
        sheetname: Name for the new worksheet.  Defaults to ``"Sheet1"``
            (or ``"Sheet2"``, ``"Sheet3"``, etc. when appending to avoid
            name collisions).
    """
    openpyxl = _require_openpyxl()

    try:
        hdrs, rows = db.select_rowsource(select_stmt)
    except ErrInfo:
        raise
    except Exception as e:
        raise ErrInfo("db", select_stmt, exception_msg=exception_desc()) from e

    # ------------------------------------------------------------------
    # Determine sheet name and open/create workbook
    # ------------------------------------------------------------------
    if append and Path(outfile).is_file():
        wb = openpyxl.load_workbook(outfile)
        existing_names = wb.sheetnames
        base = sheetname or "Sheet"
        sheet_name = base
        sheet_no = 1
        while sheet_name in existing_names:
            sheet_no += 1
            sheet_name = f"{base}{sheet_no}"
    else:
        sheet_name = sheetname or "Sheet1"
        if Path(outfile).is_file():
            filewriter_close(outfile)
            os.unlink(outfile)
        wb = openpyxl.Workbook()
        # openpyxl creates a default sheet named "Sheet"; remove it so we
        # start with a clean workbook.
        if wb.sheetnames:
            del wb[wb.sheetnames[0]]

    # ------------------------------------------------------------------
    # Ensure the inventory sheet exists
    # ------------------------------------------------------------------
    inventory_name = "Datasheets"
    if inventory_name not in wb.sheetnames:
        inv_ws = wb.create_sheet(inventory_name)
        bold_font = openpyxl.styles.Font(bold=True)
        for col_idx, hdr in enumerate(
            ("datasheet_name", "created_on", "created_by", "description", "source"),
            start=1,
        ):
            cell = inv_ws.cell(row=1, column=col_idx, value=hdr)
            cell.font = bold_font
    else:
        inv_ws = wb[inventory_name]

    # ------------------------------------------------------------------
    # Write data to a new sheet
    # ------------------------------------------------------------------
    ws = wb.create_sheet(sheet_name)
    bold_font = openpyxl.styles.Font(bold=True)

    # Header row
    for col_idx, hdr in enumerate(hdrs, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(hdr))
        cell.font = bold_font

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, item in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_cell_value(item))

    # ------------------------------------------------------------------
    # Update inventory sheet
    # ------------------------------------------------------------------
    script, lno = current_script_line()
    src = (
        f"{select_stmt} with database {_state.dbs.current().name()}, "
        f"with script {str(Path(script).resolve())}, line {lno}"
    )
    next_row = inv_ws.max_row + 1
    for col_idx, value in enumerate(
        (
            sheet_name,
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
            getpass.getuser(),
            desc,
            src,
        ),
        start=1,
    ):
        inv_ws.cell(row=next_row, column=col_idx, value=value)

    wb.save(outfile)
    wb.close()

    if _state.export_metadata is not None:
        _state.export_metadata.add(
            ExportRecord(queryname=select_stmt, outfile=outfile, zipfile=None, description=desc),
        )


def write_queries_to_xlsx(
    table_list: str,
    db: Any,
    outfile: str,
    append: bool = False,
    tee: bool = False,
    desc: str | None = None,
) -> None:
    """Write multiple tables/queries to separate worksheets in a single XLSX workbook.

    Args:
        table_list: Comma-separated list of table names (optionally schema-qualified).
        db: An execsql database adapter with a ``select_rowsource()`` method.
        outfile: Destination ``.xlsx`` file path.
        append: If ``True`` and *outfile* exists, add new sheets to the existing
            workbook rather than replacing it.
        tee: If ``True``, also pretty-print each query result to stdout.
        desc: Optional description(s).  A single string is applied to every
            sheet; a comma-separated string with the same count as *table_list*
            assigns individual descriptions per sheet.
    """
    openpyxl = _require_openpyxl()

    tables = [t.strip() for t in table_list.split(",")]
    if desc is not None:
        descriptions = [d.strip() for d in desc.split(",")]
        one_desc = len(descriptions) != len(tables)
    else:
        descriptions = []
        one_desc = False

    # ------------------------------------------------------------------
    # Open or create workbook
    # ------------------------------------------------------------------
    if Path(outfile).is_file() and not append:
        filewriter_close(outfile)
        os.unlink(outfile)

    if Path(outfile).is_file():
        wb = openpyxl.load_workbook(outfile)
    else:
        wb = openpyxl.Workbook()
        # Remove the default empty sheet created by openpyxl.
        if wb.sheetnames:
            del wb[wb.sheetnames[0]]

    # ------------------------------------------------------------------
    # Ensure the inventory sheet exists
    # ------------------------------------------------------------------
    inventory_name = "Datasheets"
    if inventory_name not in wb.sheetnames:
        inv_ws = wb.create_sheet(inventory_name)
        bold_font = openpyxl.styles.Font(bold=True)
        for col_idx, hdr in enumerate(
            ("datasheet_name", "created_on", "created_by", "description", "source"),
            start=1,
        ):
            cell = inv_ws.cell(row=1, column=col_idx, value=hdr)
            cell.font = bold_font
    else:
        inv_ws = wb[inventory_name]

    # ------------------------------------------------------------------
    # Write each table to its own sheet
    # ------------------------------------------------------------------
    bold_font = openpyxl.styles.Font(bold=True)

    for i, t in enumerate(tables):
        # Determine the table name used for the sheet label.
        if "." in t:
            st = t.split(".")
            if len(st) != 2:
                raise ErrInfo("cmd", other_msg=f"Unrecognized table specification in <{t}>")
            tblname = unquoted(st[1])
        else:
            tblname = unquoted(t)

        # Avoid duplicate sheet names.
        existing_names = wb.sheetnames
        sheet_name = tblname
        sheet_no = 1
        while sheet_name in existing_names:
            sheet_name = f"{tblname}_{sheet_no}"
            sheet_no += 1

        # Fetch data.
        select_stmt = f"select * from {t};"
        try:
            hdrs, rows = db.select_rowsource(select_stmt)
        except ErrInfo:
            raise
        except Exception as e:
            raise ErrInfo("db", select_stmt, exception_msg=exception_desc()) from e

        # Write data sheet.
        ws = wb.create_sheet(sheet_name)

        for col_idx, hdr in enumerate(hdrs, start=1):
            cell = ws.cell(row=1, column=col_idx, value=str(hdr))
            cell.font = bold_font

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, item in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=_cell_value(item))

        # Determine per-sheet description.
        if desc is None:
            d = None
        elif one_desc:
            d = desc
        else:
            d = descriptions[i]

        # Update inventory sheet.
        script, lno = current_script_line()
        src = f"From database {_state.dbs.current().name()}, with script {str(Path(script).resolve())}, line {lno}"
        next_row = inv_ws.max_row + 1
        for col_idx, value in enumerate(
            (
                sheet_name,
                datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
                getpass.getuser(),
                d,
                src,
            ),
            start=1,
        ):
            inv_ws.cell(row=next_row, column=col_idx, value=value)

        if tee and outfile.lower() != "stdout":
            prettyprint_query(select_stmt, db, "stdout", False, desc=d)

        _state.export_metadata.add(ExportRecord(queryname=select_stmt, outfile=outfile, zipfile=None, description=d))

    wb.save(outfile)
    wb.close()
