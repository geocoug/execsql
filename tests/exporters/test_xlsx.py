"""
Tests for execsql.exporters.xlsx — XLSX (Excel) export.

Tests focus on the low-level write_query_to_xlsx and write_queries_to_xlsx
functions via a lightweight stub database, and on the XlsxFile reader
(from exporters.xls) for round-trip verification.

Requires the ``openpyxl`` package (``execsql2[excel]``).  The entire module
is skipped if openpyxl is not installed.
"""

from __future__ import annotations

import datetime

import pytest

try:
    import openpyxl  # noqa: F401

    _openpyxl_available = True
except ImportError:
    _openpyxl_available = False

pytestmark = pytest.mark.skipif(
    not _openpyxl_available,
    reason="requires openpyxl (install with execsql2[excel])",
)

# ---------------------------------------------------------------------------
# Stub database
# ---------------------------------------------------------------------------


class _StubDB:
    """Minimal database stub for testing exporters without a real DB connection."""

    # Attributes required by ExportRecord.__init__
    server_name: str = "localhost"
    db_name: str = "stub_db"
    user: str | None = None

    def __init__(self, headers: list[str], rows: list[list]) -> None:
        self._headers = headers
        self._rows = rows

    def select_rowsource(self, _stmt: str):
        return self._headers[:], iter(self._rows)

    def name(self) -> str:
        return "stub_db"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _read_sheet(path: str, sheet: str | int) -> list[list]:
    """Read all rows from *sheet* in *path* using openpyxl.  Returns list of lists."""
    wb = openpyxl.load_workbook(str(path))
    if isinstance(sheet, int):
        ws = wb[wb.sheetnames[sheet - 1]]
    else:
        ws = wb[sheet]
    data = [[cell.value for cell in row] for row in ws.iter_rows()]
    wb.close()
    return data


def _sheetnames(path: str) -> list[str]:
    wb = openpyxl.load_workbook(str(path))
    names = wb.sheetnames[:]
    wb.close()
    return names


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture(autouse=True)
def _patch_state(tmp_path, monkeypatch):
    """Patch execsql.state so exporters can reference current db and script line.

    Also patch filewriter_close to a no-op: the FileWriter background process is
    not started in tests, so the real implementation would block indefinitely.
    """
    import execsql.state as _state
    import execsql.utils.fileio as _fileio

    class _FakeDB:
        server_name = "localhost"
        db_name = "test_db"
        user = None

        def name(self):
            return "test_db"

    class _FakeDbs:
        def current(self):
            return _FakeDB()

    monkeypatch.setattr(_state, "dbs", _FakeDbs())
    monkeypatch.setattr(_state, "export_metadata", _FakeMeta())
    monkeypatch.setattr(_fileio, "filewriter_close", lambda _path: None)


class _FakeMeta:
    def add(self, _record):
        pass


# ---------------------------------------------------------------------------
# Patch current_script_line for tests that call it
# ---------------------------------------------------------------------------


@pytest.fixture(autouse=True)
def _patch_script_line(monkeypatch):
    import execsql.script as _script

    monkeypatch.setattr(_script, "current_script_line", lambda: ("test_script.sql", 1))


# ---------------------------------------------------------------------------
# Tests: write_query_to_xlsx — single sheet
# ---------------------------------------------------------------------------


class TestWriteQueryToXlsx:
    def test_creates_file(self, tmp_path):
        from execsql.exporters.xlsx import write_query_to_xlsx

        out = tmp_path / "out.xlsx"
        db = _StubDB(["id", "name"], [[1, "Alice"], [2, "Bob"]])
        write_query_to_xlsx("SELECT * FROM t", db, str(out))
        assert out.exists()

    def test_header_row_is_bold(self, tmp_path):
        from execsql.exporters.xlsx import write_query_to_xlsx

        out = tmp_path / "out.xlsx"
        db = _StubDB(["col_a", "col_b"], [[10, "x"]])
        write_query_to_xlsx("SELECT * FROM t", db, str(out), sheetname="Data")
        wb = openpyxl.load_workbook(str(out))
        ws = wb["Data"]
        assert ws.cell(row=1, column=1).font.bold
        wb.close()

    def test_data_rows_written(self, tmp_path):
        from execsql.exporters.xlsx import write_query_to_xlsx

        out = tmp_path / "out.xlsx"
        db = _StubDB(["id", "name"], [[1, "Alice"], [2, "Bob"]])
        write_query_to_xlsx("SELECT * FROM t", db, str(out), sheetname="Sheet1")
        rows = _read_sheet(str(out), "Sheet1")
        # Row 0 is headers, rows 1+ are data
        assert rows[1] == [1, "Alice"]
        assert rows[2] == [2, "Bob"]

    def test_custom_sheet_name(self, tmp_path):
        from execsql.exporters.xlsx import write_query_to_xlsx

        out = tmp_path / "out.xlsx"
        db = _StubDB(["x"], [[1]])
        write_query_to_xlsx("SELECT * FROM t", db, str(out), sheetname="MyData")
        names = _sheetnames(str(out))
        assert "MyData" in names

    def test_default_sheet_name(self, tmp_path):
        from execsql.exporters.xlsx import write_query_to_xlsx

        out = tmp_path / "out.xlsx"
        db = _StubDB(["x"], [[1]])
        write_query_to_xlsx("SELECT * FROM t", db, str(out))
        names = _sheetnames(str(out))
        assert "Sheet1" in names

    def test_inventory_sheet_created(self, tmp_path):
        from execsql.exporters.xlsx import write_query_to_xlsx

        out = tmp_path / "out.xlsx"
        db = _StubDB(["x"], [[1]])
        write_query_to_xlsx("SELECT * FROM t", db, str(out))
        names = _sheetnames(str(out))
        assert "Datasheets" in names

    def test_overwrite_existing_file(self, tmp_path):
        from execsql.exporters.xlsx import write_query_to_xlsx

        out = tmp_path / "out.xlsx"
        db1 = _StubDB(["id"], [[99]])
        write_query_to_xlsx("SELECT 1", db1, str(out), sheetname="Old")

        db2 = _StubDB(["name"], [["new"]])
        write_query_to_xlsx("SELECT 2", db2, str(out), sheetname="New", append=False)

        names = _sheetnames(str(out))
        assert "Old" not in names
        assert "New" in names

    def test_append_adds_sheet(self, tmp_path):
        from execsql.exporters.xlsx import write_query_to_xlsx

        out = tmp_path / "out.xlsx"
        db = _StubDB(["x"], [[1]])
        write_query_to_xlsx("SELECT 1", db, str(out), sheetname="First")
        write_query_to_xlsx("SELECT 2", db, str(out), sheetname="Second", append=True)

        names = _sheetnames(str(out))
        assert "First" in names
        assert "Second" in names

    def test_append_deduplicates_sheet_name(self, tmp_path):
        from execsql.exporters.xlsx import write_query_to_xlsx

        out = tmp_path / "out.xlsx"
        db = _StubDB(["x"], [[1]])
        write_query_to_xlsx("SELECT 1", db, str(out), sheetname="Sheet")
        write_query_to_xlsx("SELECT 2", db, str(out), sheetname="Sheet", append=True)

        names = _sheetnames(str(out))
        # First sheet keeps name "Sheet", second becomes "Sheet2"
        assert "Sheet" in names
        assert "Sheet2" in names

    def test_empty_result_set(self, tmp_path):
        from execsql.exporters.xlsx import write_query_to_xlsx

        out = tmp_path / "out.xlsx"
        db = _StubDB(["id", "name"], [])
        write_query_to_xlsx("SELECT * FROM t", db, str(out), sheetname="Empty")
        rows = _read_sheet(str(out), "Empty")
        # Only the header row should be present
        assert len(rows) == 1
        assert rows[0] == ["id", "name"]


# ---------------------------------------------------------------------------
# Tests: type preservation
# ---------------------------------------------------------------------------


class TestXlsxTypePreservation:
    def test_integer_preserved(self, tmp_path):
        from execsql.exporters.xlsx import write_query_to_xlsx

        out = tmp_path / "types.xlsx"
        db = _StubDB(["n"], [[42]])
        write_query_to_xlsx("SELECT n FROM t", db, str(out), sheetname="T")
        rows = _read_sheet(str(out), "T")
        assert rows[1][0] == 42
        assert isinstance(rows[1][0], int)

    def test_float_preserved(self, tmp_path):
        from execsql.exporters.xlsx import write_query_to_xlsx

        out = tmp_path / "types.xlsx"
        db = _StubDB(["f"], [[3.14]])
        write_query_to_xlsx("SELECT f FROM t", db, str(out), sheetname="T")
        rows = _read_sheet(str(out), "T")
        assert abs(rows[1][0] - 3.14) < 1e-6

    def test_none_preserved(self, tmp_path):
        from execsql.exporters.xlsx import write_query_to_xlsx

        out = tmp_path / "types.xlsx"
        # Mix None with a non-None value so max_row is reliably 3 and all rows appear.
        db = _StubDB(["v"], [["present"], [None]])
        write_query_to_xlsx("SELECT v FROM t", db, str(out), sheetname="T")
        wb = openpyxl.load_workbook(str(out))
        ws = wb["T"]
        # Row 3 corresponds to the second data row (None value).
        cell_val = ws.cell(row=3, column=1).value
        wb.close()
        assert cell_val is None

    def test_bool_preserved(self, tmp_path):
        from execsql.exporters.xlsx import write_query_to_xlsx

        out = tmp_path / "types.xlsx"
        db = _StubDB(["b"], [[True], [False]])
        write_query_to_xlsx("SELECT b FROM t", db, str(out), sheetname="T")
        rows = _read_sheet(str(out), "T")
        assert rows[1][0] is True
        assert rows[2][0] is False

    def test_date_preserved(self, tmp_path):
        from execsql.exporters.xlsx import write_query_to_xlsx

        out = tmp_path / "types.xlsx"
        d = datetime.date(2024, 6, 15)
        db = _StubDB(["d"], [[d]])
        write_query_to_xlsx("SELECT d FROM t", db, str(out), sheetname="T")
        rows = _read_sheet(str(out), "T")
        # openpyxl reads dates back as datetime.datetime objects
        val = rows[1][0]
        assert isinstance(val, datetime.datetime | datetime.date)
        assert val.year == 2024
        assert val.month == 6
        assert val.day == 15

    def test_datetime_preserved(self, tmp_path):
        from execsql.exporters.xlsx import write_query_to_xlsx

        out = tmp_path / "types.xlsx"
        dt = datetime.datetime(2024, 3, 10, 14, 30, 0)
        db = _StubDB(["ts"], [[dt]])
        write_query_to_xlsx("SELECT ts FROM t", db, str(out), sheetname="T")
        rows = _read_sheet(str(out), "T")
        val = rows[1][0]
        assert isinstance(val, datetime.datetime)
        assert val.year == 2024
        assert val.hour == 14

    def test_string_preserved(self, tmp_path):
        from execsql.exporters.xlsx import write_query_to_xlsx

        out = tmp_path / "types.xlsx"
        db = _StubDB(["s"], [["hello world"]])
        write_query_to_xlsx("SELECT s FROM t", db, str(out), sheetname="T")
        rows = _read_sheet(str(out), "T")
        assert rows[1][0] == "hello world"

    def test_time_stored_as_string(self, tmp_path):
        from execsql.exporters.xlsx import write_query_to_xlsx

        out = tmp_path / "types.xlsx"
        t = datetime.time(9, 5, 3)
        db = _StubDB(["t"], [[t]])
        write_query_to_xlsx("SELECT t FROM t", db, str(out), sheetname="T")
        rows = _read_sheet(str(out), "T")
        # time is stored as HH:MM:SS string
        assert rows[1][0] == "09:05:03"


# ---------------------------------------------------------------------------
# Tests: write_queries_to_xlsx — multi-sheet
# ---------------------------------------------------------------------------


class TestWriteQueriesToXlsx:
    def test_creates_multiple_sheets(self, tmp_path):
        from execsql.exporters.xlsx import write_queries_to_xlsx

        out = tmp_path / "multi.xlsx"
        db = _StubDB(["id", "val"], [[1, "a"], [2, "b"]])
        write_queries_to_xlsx("t1, t2", db, str(out))

        names = _sheetnames(str(out))
        assert "t1" in names
        assert "t2" in names

    def test_each_sheet_has_data(self, tmp_path):
        from execsql.exporters.xlsx import write_queries_to_xlsx

        out = tmp_path / "multi.xlsx"
        db = _StubDB(["id", "val"], [[1, "x"]])
        write_queries_to_xlsx("tbl_a, tbl_b", db, str(out))

        rows_a = _read_sheet(str(out), "tbl_a")
        assert rows_a[0] == ["id", "val"]
        assert rows_a[1] == [1, "x"]

    def test_inventory_sheet_present(self, tmp_path):
        from execsql.exporters.xlsx import write_queries_to_xlsx

        out = tmp_path / "multi.xlsx"
        db = _StubDB(["x"], [[1]])
        write_queries_to_xlsx("alpha", db, str(out))

        assert "Datasheets" in _sheetnames(str(out))

    def test_append_mode_adds_sheets(self, tmp_path):
        from execsql.exporters.xlsx import write_queries_to_xlsx

        out = tmp_path / "multi.xlsx"
        db = _StubDB(["x"], [[1]])
        write_queries_to_xlsx("t1", db, str(out))
        write_queries_to_xlsx("t2", db, str(out), append=True)

        names = _sheetnames(str(out))
        assert "t1" in names
        assert "t2" in names

    def test_non_append_replaces_file(self, tmp_path):
        from execsql.exporters.xlsx import write_queries_to_xlsx

        out = tmp_path / "multi.xlsx"
        db = _StubDB(["x"], [[1]])
        write_queries_to_xlsx("t1", db, str(out))
        write_queries_to_xlsx("t2", db, str(out), append=False)

        names = _sheetnames(str(out))
        assert "t1" not in names
        assert "t2" in names

    def test_duplicate_table_names_deduped(self, tmp_path):
        from execsql.exporters.xlsx import write_queries_to_xlsx

        out = tmp_path / "multi.xlsx"
        db = _StubDB(["x"], [[1]])
        write_queries_to_xlsx("same, same", db, str(out))

        names = _sheetnames(str(out))
        assert "same" in names
        assert "same_1" in names

    def test_schema_qualified_table_uses_table_part_as_sheet_name(self, tmp_path):
        from execsql.exporters.xlsx import write_queries_to_xlsx

        out = tmp_path / "multi.xlsx"
        db = _StubDB(["col"], [["v"]])
        write_queries_to_xlsx("public.mytable", db, str(out))

        names = _sheetnames(str(out))
        assert "mytable" in names
